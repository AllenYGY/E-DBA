from typing import Any, Dict
from io import BytesIO

from fastapi import APIRouter, Body, Depends, HTTPException, status, UploadFile, File
from fastapi.encoders import jsonable_encoder
from sqlalchemy.orm import Session
import openpyxl

from app import schemas, crud
from app.api import deps
from app.models import User, Organization
from app.models.log import LogType
from app.models.enums import UserRole

router = APIRouter()

def can_create(creator_role, target_role):
    if creator_role == UserRole.T_ADMIN:
        return True
    if creator_role == UserRole.SENIOR_E_ADMIN:
        return target_role in [UserRole.E_ADMIN, UserRole.O_CONVENER, UserRole.DATA_USER]
    if creator_role == UserRole.E_ADMIN:
        return target_role in [UserRole.O_CONVENER, UserRole.DATA_USER]
    if creator_role == UserRole.O_CONVENER:
        return target_role in [UserRole.DATA_USER]
    return False


@router.post("/", response_model=schemas.User)
async def create_user(
    *,
    db: Session = Depends(deps.get_db),
    user_in: schemas.UserCreate,
    current_user: User = Depends(deps.get_current_o_convener),
) -> Any:
    """
    Create a new user. Need to be one of the following roles: (E-Admin, T-Admin, Senior E-Admin, O-Convener)
    """
    user = crud.user.get_by_email(db, email=user_in.email)
    if user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The email has been registered",
        )
    target_role = user_in.role if isinstance(user_in.role, str) else user_in.role.value
    creator_role = current_user.role if isinstance(current_user.role, str) else current_user.role.value
    if not can_create(creator_role, target_role):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=f"{creator_role} does not have permission to create {target_role} user",
        )
    user_in.organization_id = current_user.organization_id
    user = crud.user.create(db, obj_in=user_in)
    
    crud.log.create_log(
        db=db,
        user_id=current_user.id,
        log_type=LogType.SYSTEM,
        organization_id=current_user.organization_id,
        action="Create user",
        details=f"User {current_user.email} created user {user.email}"
    )
    return user


@router.get("/me")
async def read_user_me(
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
) -> Any:
    """获取当前登录用户信息和组织"""
    organization = None
    if current_user.organization_id:
        organization = db.query(Organization).filter(Organization.id == current_user.organization_id).first()
    
    user_info = {
        "id": current_user.id,
        "email": current_user.email,
        "username": current_user.username,
        "role": current_user.role,
        "is_active": current_user.is_active,
        "is_deleted": current_user.is_deleted,
        "created_at": current_user.created_at,
        "updated_at": current_user.updated_at,
        "deleted_at": current_user.deleted_at,
        "organization_id": current_user.organization_id,
        "permission_level": current_user.permission_level,
        "balance": current_user.balance,
        "organization_short_name": organization.name if organization else None,
        "organization_full_name": organization.full_name if organization else None
    }
    return user_info


@router.put("/me", response_model=schemas.User)
async def update_user_me(
    *,
    db: Session = Depends(deps.get_db),
    email: str = Body(None),
    username: str = Body(None),
    password: str = Body(None),
    current_user: User = Depends(deps.get_current_active_user),
) -> Any:
    """
    Update current user information
    """
    current_user_data = jsonable_encoder(current_user)
    user_in = schemas.UserUpdate(**current_user_data)
    
    if email is not None:
        user_in.email = email
    if password is not None:
        user_in.password = password
    if username is not None:
        user_in.username = username
        
    # 检查邮箱是否已注册 除了当前用户
    if user_in.email and user_in.email != current_user.email:
        existing_user = crud.user.get_by_email(db, email=user_in.email)
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The email has been registered",
            )
        
    user = crud.user.update(db, db_obj=current_user, obj_in=user_in)
    
    # Record user update log
    crud.log.create_log(
        db=db,
        user_id=current_user.id,
        organization_id=current_user.organization_id,
        log_type=LogType.SYSTEM,
        action="Update personal information",
        details=f"User {user.email} updated personal information"
    )
    
    return user


@router.get("/search", response_model=Dict[str, Any])
async def search_users(
    *,
    db: Session = Depends(deps.get_db),
    username: str = None,
    email: str = None,
    role: str = None,
    organization_id: int = None,
    permission_level: int = None,
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(deps.get_current_o_convener),
) -> Any:
    """
    Search users with multiple conditions
    Need to be one of the following roles: (E-Admin, T-Admin, Senior E-Admin, O-Convener)
    """
    query = db.query(User)
    if username:
        query = query.filter(User.username.ilike(f"%{username}%"))
    if email:
        query = query.filter(User.email.ilike(f"%{email}%"))
    if role:
        query = query.filter(User.role == role)
    if organization_id is not None:
        query = query.filter(User.organization_id == organization_id)
    if permission_level is not None:
        query = query.filter(User.permission_level == permission_level)
    total = query.count()
    users = query.offset(skip).limit(limit).all()
    crud.log.create_log(
        db=db,
        user_id=current_user.id,
        organization_id=current_user.organization_id,
        log_type=LogType.SYSTEM,
        action="Search users",
        details=f"Admin {current_user.email} searched users with multiple conditions"
    )
    return {
        "items": [schemas.User.from_orm(user) for user in users],
        "total": total
    }

@router.get("/{user_id}", response_model=schemas.User)
async def read_user_by_id(
    user_id: int,
    current_user: User = Depends(deps.get_current_active_user),
    db: Session = Depends(deps.get_db),
) -> Any:
    """
    Get user information by ID
    Need to be one of the following roles: (E-Admin, T-Admin, Senior E-Admin, O-Convener)
    """
    user = crud.user.get(db, id=user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    if user == current_user:
        return user
    if current_user.role not in [UserRole.E_ADMIN, UserRole.T_ADMIN, UserRole.SENIOR_E_ADMIN]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Insufficient permissions, cannot access other user information",
        )
    
    # Record log of viewing user information
    crud.log.create_log(
        db=db,
        user_id=current_user.id,
        organization_id=current_user.organization_id,
        log_type=LogType.SYSTEM,
        action="View user information",
        details=f"Admin {current_user.email} viewed user {user.email} information"
    )
    
    return user

@router.put("/{user_id}", response_model=schemas.User)
async def update_user(
    *,
    db: Session = Depends(deps.get_db),
    user_id: int,
    user_in: schemas.UserUpdate,
    current_user: User = Depends(deps.get_current_o_convener),
) -> Any:
    """
    Update user information
    Need to be one of the following roles: (E-Admin, T-Admin, Senior E-Admin, O-Convener)
    """
    user = crud.user.get(db, id=user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found, cannot update",
        )
    
    # Check if the email is repeated
    if user_in.email and user_in.email != user.email:
        existing_user = crud.user.get_by_email(db, email=user_in.email)
        # Only report an error if the found user is not the current user and the email is different
        if existing_user and existing_user.id != user.id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The email has been registered",
            )
        
    if current_user.role == UserRole.O_CONVENER and user_in.organization_id != current_user.organization_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="O-Convener can only update user information within their own organization",
        )

    user = crud.user.update(db, db_obj=user, obj_in=user_in)
    
    # Record user update log
    crud.log.create_log(
        db=db,
        user_id=current_user.id,
        organization_id=current_user.organization_id,
        log_type=LogType.SYSTEM,
        action="Update user information",
        details=f"Admin {current_user.email} updated user {user.email} information"
    )
    return user


@router.delete("/{user_id}", response_model=schemas.User)
async def delete_user(
    *,
    db: Session = Depends(deps.get_db),
    user_id: int,
    current_user: User = Depends(deps.get_current_o_convener),
) -> Any:
    """
    Delete user
    Need to be one of the following roles: (T-Admin, Senior E-Admin, E-Admin, O-Convener)
    """
    user = crud.user.get(db, id=user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found, cannot delete",
        )
    if user.id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete your own account",
        )
    
    if current_user.role == UserRole.O_CONVENER and user.organization_id != current_user.organization_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="O-Convener can only delete user information within their own organization",
        )
    
    user = crud.user.remove(db, id=user_id)
    
    # Record user deletion log
    crud.log.create_log(
        db=db,
        user_id=current_user.id,
        organization_id=current_user.organization_id,
        log_type=LogType.SYSTEM,
        action="Delete user",
        details=f"{current_user.email} deleted user {user.email}"
    )
    
    return user

@router.post("/{user_id}/edit_balance", response_model=dict)
async def edit_balance(
    user_id: int,
    *,
    db: Session = Depends(deps.get_db),
    amount: float = Body(...),
    current_user: User = Depends(deps.get_current_active_user),
) -> Any:
    """
    Add or deduct user balance
    Need to be one of the following roles: (E-Admin, T-Admin, Senior E-Admin, O-Convener)
    
    When deducting money (amount < 0), the balance after deduction must be >= 0
    """
    user = crud.user.get(db, id=user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found",
        )
    
    # Check if balance is enough when deducting money
    if amount < 0 and user.balance + amount < 0:
        return {
            "message": "Balance not enough",
            "user_id": user.id,
            "amount": amount,
            "current_balance": user.balance
        }
    
    # Edit user balance
    new_balance = user.balance + amount
    user = crud.user.update(db, db_obj=user, obj_in={"balance": new_balance})
    
    # Record log
    action_type = "Deduct" if amount < 0 else "Add"
    crud.log.create_log(
        db=db,
        user_id=current_user.id,
        organization_id=current_user.organization_id,
        log_type=LogType.PAYMENT,
        action=f"{action_type} user balance",
        details=f"Admin {current_user.email} {action_type.lower()}ed user {user.email} balance by {abs(amount)} yuan, current balance: {user.balance} yuan"
    )
    
    return {
        "message": f"{action_type} user balance successfully",
        "user_id": user.id,
        "amount": amount,
        "current_balance": user.balance
    }

@router.post("/batch_create", response_model=dict)
async def batch_create_users(
    *,
    db: Session = Depends(deps.get_db),
    file: UploadFile = File(...),
    current_user: User = Depends(deps.get_current_o_convener),
) -> Any:
    """
    批量导入用户（Excel），只支持 .xlsx
    """
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    content = await file.read()
    excel_io = BytesIO(content)
    wb = openpyxl.load_workbook(excel_io, read_only=True)
    ws = wb.active

    # 解析表头，全部转小写并去除空格
    header = [str(cell.value).strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print("header", header)
    required_fields = {"username", "email", "password"}
    if not required_fields.issubset(set(header)):
        raise HTTPException(status_code=400, detail=f"Excel must contain columns: {', '.join(required_fields)}")

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(header, row))
        # 权限限制
        role = data.get("role", UserRole.DATA_USER)
        if current_user.role == UserRole.O_CONVENER and role != UserRole.DATA_USER:
            results.append({"email": data.get("email"), "success": False, "reason": "O-Convener can only create Data User"})
            continue
        # 校验邮箱是否已注册
        if crud.user.get_by_email(db, email=data["email"]):
            results.append({"email": data["email"], "success": False, "reason": "Email already registered"})
            continue
        # 构造 UserCreate
        try:
            user_in = schemas.UserCreate(
                username=data["username"],
                email=data["email"],
                password=data["password"],
                permission_level=int(data.get("permission_level", 1)),
                role=role,
                organization_id=current_user.organization_id
            )
            user = crud.user.create(db, obj_in=user_in)
            results.append({"email": data.get("email"), "success": True, "user_id": user.id, "username": user.username})
        except Exception as e:
            results.append({"email": data.get("email"), "success": False, "reason": str(e)})

    crud.log.create_log(
        db=db,
        user_id=current_user.id,
        log_type=LogType.SYSTEM,
        organization_id=current_user.organization_id,
        action="Batch create users",
        details=f"User {current_user.email} batch created users via Excel"
    )
    return {"results": results}